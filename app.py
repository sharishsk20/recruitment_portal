"""
TÜV SÜD Recruitment Portal — Flask backend
Run:  pip install -r requirements.txt  then  python app.py
Open: http://127.0.0.1:5000
"""
import io
import sqlite3
from flask import Flask, request, jsonify, render_template, send_file, g
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
DB = "candidates.db"

# TÜV SÜD palette (deep blue + grey + white)
TUV_BLUE = "005AA0"
TUV_BLUE_DARK = "003A70"
TUV_GREY = "53565A"

FIELDS = ["name", "role", "location", "experience", "certifications",
          "iso_certified", "source", "linkedin_sourced", "stage", "status", "notes"]


def db():
    if "db" not in g:
        g.db = sqlite3.connect(DB)
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(exc):
    d = g.pop("db", None)
    if d is not None:
        d.close()


def init_db():
    con = sqlite3.connect(DB)
    con.execute("""CREATE TABLE IF NOT EXISTS candidates(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        role TEXT, location TEXT, experience REAL,
        certifications TEXT, iso_certified TEXT,
        source TEXT, linkedin_sourced TEXT,
        stage TEXT, status TEXT, notes TEXT)""")
    con.execute("""CREATE TABLE IF NOT EXISTS positions(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        position TEXT NOT NULL,
        location TEXT,
        status TEXT,
        date_opened TEXT,
        date_closed TEXT,
        days TEXT,
        sf_applications TEXT,
        linkedin_views TEXT,
        linkedin_applications TEXT,
        recruiter_experience TEXT,
        notes TEXT)""")
    con.commit()
    con.close()


init_db()


def query(filters):
    sql = "SELECT * FROM candidates WHERE 1=1"
    params = []
    if filters.get("text"):
        sql += " AND (LOWER(name) LIKE ? OR LOWER(certifications) LIKE ?)"
        t = f"%{filters['text'].lower()}%"
        params += [t, t]
    if filters.get("role"):
        sql += " AND role=?"; params.append(filters["role"])
    if filters.get("iso"):
        sql += " AND iso_certified=?"; params.append(filters["iso"])
    if filters.get("source"):
        sql += " AND source=?"; params.append(filters["source"])
    if filters.get("linkedin_sourced"):
        sql += " AND linkedin_sourced=?"; params.append(filters["linkedin_sourced"])
    if filters.get("status"):
        sql += " AND status=?"; params.append(filters["status"])
    if filters.get("min_exp"):
        try:
            sql += " AND experience>=?"; params.append(float(filters["min_exp"]))
        except ValueError:
            pass
    sql += " ORDER BY id DESC"
    return db().execute(sql, params).fetchall()


# ---------------------------------------------------------------------------
# Rule-based natural-language / free-text candidate parser.
# No external API is used — this is regex + keyword matching, so it works
# offline but is NOT true language understanding. Always review parsed
# results before saving (the frontend enforces a preview step).
# ---------------------------------------------------------------------------
import re

KNOWN_ROLES = ["Manager – Inside Sales", "Manager – Sales", "QSA"]
KNOWN_LOCATIONS = ["Gurgaon", "Gurugram", "Mumbai", "Bangalore", "Bengaluru", "Delhi",
                   "Chennai", "Pune", "Hyderabad", "Noida", "Kolkata", "Ahmedabad",
                   "Jaipur", "Chandigarh", "Kochi", "Cochin", "Coimbatore", "Indore",
                   "Lucknow", "Bhopal", "Nagpur", "Vadodara", "Remote"]

# Words that can never be a candidate name, used to keep the name-extraction
# heuristic below from mistaking a role/location/keyword phrase for a name.
_NAME_STOPWORDS = {
    "manager", "inside", "sales", "qsa", "iso", "linkedin", "naukri", "sf",
    "referral", "referred", "joined", "rejected", "declined", "hold", "onhold",
    "screening", "final", "round", "first", "1st", "2nd", "offer", "stage",
    "years", "year", "yrs", "yr", "yoe", "lead", "auditor", "fresher", "exp",
    "experience", "shortlisted", "selected", "technical", "hr", "walkin",
    "indeed", "monster", "employee", "internal", "reference", "candidate",
    "profile", "resume", "applied", "application",
} | {loc.lower() for loc in KNOWN_LOCATIONS}


def _empty_candidate():
    return {"name": "", "role": "Other", "location": "", "experience": 0,
            "certifications": "None listed", "iso_certified": "No", "source": "Other",
            "linkedin_sourced": "No", "stage": "—", "status": "In Process", "notes": ""}


def _extract_name(line, chunks):
    """Best-effort candidate name extraction.

    Preferred signal: a run of 2-4 Title-Case words (e.g. "Rohit Verma"),
    since that's what a human name looks like regardless of where it sits
    in the line. Falls back to the old "first non-keyword chunk" heuristic
    for lines that are all lowercase or otherwise don't have a clean
    Title-Case run.
    """
    for m in re.finditer(r'\b[A-Z][a-zA-Z.]+(?:\s+[A-Z][a-zA-Z.]+){1,3}\b', line):
        words = m.group(0).split()
        if all(w.strip('.').lower() not in _NAME_STOPWORDS for w in words) \
           and not any(m.group(0).lower() == r.lower() for r in KNOWN_ROLES):
            return m.group(0)

    skip_pattern = re.compile(
        r'^(iso|linkedin|naukri|sf|referral|joined|reject|declin|hold|screening|'
        r'final round|1st round|first round|offer|years?|yrs?|\d+(\.\d+)?\s*(years?|yrs?)?)$',
        re.I)
    for c in chunks:
        if skip_pattern.match(c) or any(r.lower() == c.lower() for r in KNOWN_ROLES) \
           or any(loc.lower() == c.lower() for loc in KNOWN_LOCATIONS):
            continue
        if re.search(r'\d', c) and len(c) < 6:
            continue
        return c
    return chunks[0] if chunks else "Unnamed Candidate"


def parse_candidate_line(line):
    """Extract structured fields from one line of free text using heuristics."""
    result = _empty_candidate()
    low = line.lower()

    # Role
    for r in KNOWN_ROLES:
        if r.lower() in low:
            result["role"] = r
            break
    else:
        if "inside sales" in low:
            result["role"] = "Manager – Inside Sales"
        elif "sales" in low:
            result["role"] = "Manager – Sales"
        elif re.search(r'\bqsa\b|lead auditor', low):
            result["role"] = "QSA"

    # Location
    for loc in KNOWN_LOCATIONS:
        if re.search(r'\b' + re.escape(loc.lower()) + r'\b', low):
            result["location"] = "Bangalore" if loc == "Bengaluru" else \
                ("Gurgaon" if loc == "Gurugram" else ("Kochi" if loc == "Cochin" else loc))
            break

    # Experience (years) — handles "5 years", "5+ yrs", "3-5 years" (takes
    # the lower bound), "fresher"/"entry level" (0), and bare "X yoe"
    if re.search(r'\bfresher\b|\bentry[\s-]level\b|\b0\s*(?:years?|yrs?)\b', low):
        result["experience"] = 0
    else:
        m = re.search(r'(\d+(?:\.\d+)?)\s*(?:-|to)\s*\d+(?:\.\d+)?\s*(?:years?|yrs?|yoe)\b', line, re.I)
        if not m:
            m = re.search(r'(\d+(?:\.\d+)?)\s*\+?\s*(?:years?|yrs?|yoe)\b', line, re.I)
        if m:
            result["experience"] = float(m.group(1))

    # ISO certifications (and other common industry certs, kept in the same
    # field since the DB has no separate column for non-ISO certs)
    iso_matches = re.findall(r'ISO\s?\d{3,6}(?:[:\-]?\d{0,4})?(?:\s+(?:lead\s+auditor|qsa))?', line, re.I)
    other_certs = re.findall(r'\b(CISA|CISSP|CISM|PMP|Six\s?Sigma(?:\s+(?:Green|Black)\s+Belt)?|CEH)\b', line, re.I)
    seen = []
    for im in iso_matches + other_certs:
        v = re.sub(r'\s+', ' ', im.strip())
        if v.lower() not in [s.lower() for s in seen]:
            seen.append(v)
    if seen:
        result["certifications"] = ", ".join(seen)
        result["iso_certified"] = "Yes" if iso_matches else result["iso_certified"]

    # Source / LinkedIn
    if "linkedin" in low:
        result["source"] = "LinkedIn"
        result["linkedin_sourced"] = "Yes"
    elif re.search(r'\bsf\b|naukri|successfactors|indeed|monster', low):
        result["source"] = "Naukri/SF"
    elif re.search(r'referral|referred|employee reference|internal reference', low):
        result["source"] = "Referral"

    # Interview stage
    if "final round" in low or re.search(r'\bl3\b', low):
        result["stage"] = "Final Round"
    elif re.search(r'1st round|first round|technical round|tech round|\bl1\b', low):
        result["stage"] = "1st Round Interview"
    elif re.search(r'\boffer\b|selected', low):
        result["stage"] = "Offer Stage"
    elif re.search(r'screening|shortlisted|hr round|\bl2\b', low):
        result["stage"] = "Screening"

    # Status
    if "joined" in low:
        result["status"] = "Joined"
    elif re.search(r'reject', low):
        result["status"] = "Rejected"
    elif re.search(r'declin|withdr|backed out', low):
        result["status"] = "Offer Declined"
    elif "hold" in low:
        result["status"] = "On Hold"
    else:
        result["status"] = "In Process"

    # Name
    chunks = [c.strip() for c in re.split(r'[-–—|,;]', line) if c.strip()]
    result["name"] = _extract_name(line, chunks)
    result["_raw"] = line
    return result


def parse_candidates_text(text):
    lines = [l.strip() for l in text.strip().split("\n") if l.strip()]
    return [parse_candidate_line(l) for l in lines]


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/candidates", methods=["GET"])
def list_candidates():
    filters = {
        "text": request.args.get("text", ""),
        "role": request.args.get("role", ""),
        "iso": request.args.get("iso", ""),
        "source": request.args.get("source", ""),
        "status": request.args.get("status", ""),
        "linkedin_sourced": request.args.get("linkedin_sourced", ""),
        "min_exp": request.args.get("min_exp", ""),
    }
    rows = [dict(r) for r in query(filters)]
    return jsonify(rows)


@app.route("/api/candidates", methods=["POST"])
def add_candidate():
    d = request.get_json(force=True)
    if not d.get("name", "").strip():
        return jsonify({"error": "Name is required"}), 400
    db().execute(
        """INSERT INTO candidates
        (name,role,location,experience,certifications,iso_certified,source,linkedin_sourced,stage,status,notes)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
        (d.get("name", "").strip(), d.get("role", ""), d.get("location", ""),
         float(d.get("experience") or 0), d.get("certifications", "") or "None listed",
         d.get("iso_certified", "No"), d.get("source", ""), d.get("linkedin_sourced", "No"),
         d.get("stage", ""), d.get("status", "In Process"), d.get("notes", "")))
    db().commit()
    return jsonify({"ok": True})


@app.route("/api/candidates/<int:cid>", methods=["PUT"])
def update_candidate(cid):
    d = request.get_json(force=True)
    if not d.get("name", "").strip():
        return jsonify({"error": "Name is required"}), 400
    db().execute(
        """UPDATE candidates SET name=?, role=?, location=?, experience=?, certifications=?,
           iso_certified=?, source=?, linkedin_sourced=?, stage=?, status=?, notes=? WHERE id=?""",
        (d.get("name", "").strip(), d.get("role", ""), d.get("location", ""),
         float(d.get("experience") or 0), d.get("certifications", "") or "None listed",
         d.get("iso_certified", "No"), d.get("source", ""), d.get("linkedin_sourced", "No"),
         d.get("stage", ""), d.get("status", "In Process"), d.get("notes", ""), cid))
    db().commit()
    return jsonify({"ok": True})


@app.route("/api/candidates/<int:cid>", methods=["DELETE"])
def delete_candidate(cid):
    db().execute("DELETE FROM candidates WHERE id=?", (cid,))
    db().commit()
    return jsonify({"ok": True})


@app.route("/api/parse_nlp", methods=["POST"])
def parse_nlp():
    d = request.get_json(force=True)
    text = d.get("text", "")
    if not text.strip():
        return jsonify({"error": "No text provided"}), 400
    parsed = parse_candidates_text(text)
    return jsonify(parsed)


@app.route("/api/import_file", methods=["POST"])
def import_file():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    filename = (f.filename or "").lower()

    # Flexible header matching: normalize header text to a canonical field name
    def norm(h):
        h = (h or "").strip().lower()
        h = re.sub(r'[^a-z0-9]+', ' ', h).strip()
        return h

    header_map = {
        "name": "name", "candidate name": "name",
        "role": "role", "role applied for": "role", "position": "role",
        "location": "location", "loc": "location", "city": "location",
        "experience": "experience", "experience yrs": "experience", "exp": "experience",
        "certifications": "certifications", "certification": "certifications",
        "iso certified": "iso_certified", "iso": "iso_certified",
        "source": "source",
        "linkedin sourced": "linkedin_sourced", "linkedin": "linkedin_sourced",
        "interview stage": "stage", "stage": "stage",
        "status": "status",
        "recruiter notes": "notes", "notes": "notes",
    }

    rows = []
    try:
        if filename.endswith(".csv"):
            import csv as csv_mod
            text = f.stream.read().decode("utf-8", errors="ignore")
            reader = csv_mod.DictReader(io.StringIO(text))
            raw_rows = list(reader)
            headers = reader.fieldnames or []
        else:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(f.read()), data_only=True)
            ws = wb.active
            all_vals = list(ws.iter_rows(values_only=True))
            # find the first row that looks like a header (matches known field names)
            header_row_idx = 0
            for i, r in enumerate(all_vals[:10]):
                normed = [norm(str(v)) for v in r if v is not None]
                if sum(1 for n in normed if n in header_map) >= 2:
                    header_row_idx = i
                    break
            headers = [str(v) if v is not None else "" for v in all_vals[header_row_idx]]
            raw_rows = []
            for r in all_vals[header_row_idx + 1:]:
                if all(v is None for v in r):
                    continue
                raw_rows.append({headers[i]: r[i] for i in range(len(headers)) if i < len(r)})

        col_lookup = {h: header_map.get(norm(h)) for h in headers}

        for raw in raw_rows:
            c = _empty_candidate()
            for h, v in raw.items():
                field = col_lookup.get(h)
                if not field or v is None:
                    continue
                v = str(v).strip()
                if field == "experience":
                    try:
                        c["experience"] = float(re.sub(r'[^\d.]', '', v) or 0)
                    except ValueError:
                        pass
                elif field in ("iso_certified", "linkedin_sourced"):
                    c[field] = "Yes" if v.lower() in ("yes", "y", "true", "1") else "No"
                else:
                    c[field] = v
            if c["certifications"] and c["certifications"] != "None listed" and "iso" in c["certifications"].lower():
                c["iso_certified"] = "Yes"
            if c["name"]:
                rows.append(c)
    except Exception as e:
        return jsonify({"error": f"Could not parse file: {e}"}), 400

    if not rows:
        return jsonify({"error": "No recognizable candidate rows found. Check column headers."}), 400
    return jsonify(rows)


@app.route("/api/bulk_add", methods=["POST"])
def bulk_add():
    d = request.get_json(force=True)
    candidates = d.get("candidates", [])
    added = 0
    for c in candidates:
        name = (c.get("name") or "").strip()
        if not name:
            continue
        db().execute(
            """INSERT INTO candidates
            (name,role,location,experience,certifications,iso_certified,source,linkedin_sourced,stage,status,notes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (name, c.get("role", ""), c.get("location", ""),
             float(c.get("experience") or 0), c.get("certifications", "") or "None listed",
             c.get("iso_certified", "No"), c.get("source", ""), c.get("linkedin_sourced", "No"),
             c.get("stage", ""), c.get("status", "In Process"), c.get("notes", "")))
        added += 1
    db().commit()
    return jsonify({"ok": True, "added": added})


def query_positions():
    return db().execute("SELECT * FROM positions ORDER BY id").fetchall()


@app.route("/api/positions", methods=["GET"])
def list_positions():
    return jsonify([dict(r) for r in query_positions()])


@app.route("/api/positions", methods=["POST"])
def add_position():
    d = request.get_json(force=True)
    if not d.get("position", "").strip():
        return jsonify({"error": "Position title is required"}), 400
    db().execute(
        """INSERT INTO positions
        (position,location,status,date_opened,date_closed,days,sf_applications,
         linkedin_views,linkedin_applications,recruiter_experience,notes)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
        (d.get("position", "").strip(), d.get("location", ""), d.get("status", ""),
         d.get("date_opened", ""), d.get("date_closed", ""), d.get("days", ""),
         d.get("sf_applications", ""), d.get("linkedin_views", ""),
         d.get("linkedin_applications", ""), d.get("recruiter_experience", ""), d.get("notes", "")))
    db().commit()
    return jsonify({"ok": True})


@app.route("/api/positions/<int:pid>", methods=["PUT"])
def update_position(pid):
    d = request.get_json(force=True)
    if not d.get("position", "").strip():
        return jsonify({"error": "Position title is required"}), 400
    db().execute(
        """UPDATE positions SET position=?, location=?, status=?, date_opened=?, date_closed=?,
           days=?, sf_applications=?, linkedin_views=?, linkedin_applications=?,
           recruiter_experience=?, notes=? WHERE id=?""",
        (d.get("position", "").strip(), d.get("location", ""), d.get("status", ""),
         d.get("date_opened", ""), d.get("date_closed", ""), d.get("days", ""),
         d.get("sf_applications", ""), d.get("linkedin_views", ""),
         d.get("linkedin_applications", ""), d.get("recruiter_experience", ""), d.get("notes", ""), pid))
    db().commit()
    return jsonify({"ok": True})


@app.route("/api/positions/<int:pid>", methods=["DELETE"])
def delete_position(pid):
    db().execute("DELETE FROM positions WHERE id=?", (pid,))
    db().commit()
    return jsonify({"ok": True})


@app.route("/api/seed_positions", methods=["POST"])
def seed_positions():
    samples = [
        ("Manager – Sales", "Gurgaon", "Yet to Join", "Oct 2025", "Offer closed – May 2026",
         "~7 months (approx.)", "47", "Not tracked on LinkedIn", "N/A", "Not specified",
         "Sourced entirely outside LinkedIn. Role open since Oct 2025; offer accepted in May 2026, joining still pending."),
        ("Manager – Inside Sales", "Mumbai", "Joined", "Feb 2026", "11-May-2026",
         "77", "49", "1483", "266", "Smooth — high LinkedIn traction",
         "Strongest LinkedIn funnel of the three roles (1,483 views → 266 applications). Closed and joined within 77 days of opening."),
        ("QSA", "Mumbai", "Interviews in Progress", "10-Feb-2026", "Open (1st round interviews in progress)",
         "79", "24", "1421", "123", "Initially difficult; improved to moderate after sourcing campaign",
         "~30 LinkedIn connects and 40+ DMs sent (effort shared across this and other open roles). Direct LinkedIn outreach generated strong referrals and candidate interest."),
    ]
    db().executemany(
        """INSERT INTO positions
        (position,location,status,date_opened,date_closed,days,sf_applications,
         linkedin_views,linkedin_applications,recruiter_experience,notes)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)""", samples)
    db().commit()
    return jsonify({"ok": True})


@app.route("/api/clear_positions", methods=["POST"])
def clear_positions():
    db().execute("DELETE FROM positions")
    db().commit()
    return jsonify({"ok": True})


@app.route("/api/seed", methods=["POST"])
def seed():
    samples = [
        ("[Sample] Aarti Sharma", "QSA", "Mumbai", 7, "ISO 9001 QSA, ISO 27001", "Yes", "LinkedIn", "Yes", "1st Round Interview", "In Process", "Strong referral"),
        ("[Sample] Rohit Verma", "Manager – Inside Sales", "Mumbai", 6, "ISO 9001", "Yes", "LinkedIn", "Yes", "Offer Stage", "Joined", ""),
        ("[Sample] Neha Gupta", "Manager – Sales", "Gurgaon", 8, "ISO 9001 Lead Auditor", "Yes", "Naukri/SF", "No", "Final Round", "In Process", ""),
        ("[Sample] Karan Mehta", "QSA", "Mumbai", 3, "None listed", "No", "Naukri/SF", "No", "Screening", "Rejected", "No ISO cert"),
    ]
    db().executemany(
        """INSERT INTO candidates
        (name,role,location,experience,certifications,iso_certified,source,linkedin_sourced,stage,status,notes)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)""", samples)
    db().commit()
    return jsonify({"ok": True})


@app.route("/api/clear", methods=["POST"])
def clear():
    db().execute("DELETE FROM candidates")
    db().commit()
    return jsonify({"ok": True})


from openpyxl.utils import get_column_letter as _col


def _to_number(v):
    """Best-effort numeric parse; returns None if not cleanly numeric."""
    try:
        return float(re.sub(r'[^\d.]', '', str(v)))
    except (ValueError, TypeError):
        return None


def build_hiring_tracker(ws, white, blue_fill, grey_fill, border, position_rows):
    ws.title = "Hiring Tracker"
    ws.sheet_view.showGridLines = False

    # Title banner
    ws.merge_cells("A1:L1")
    ws["A1"] = "TÜV SÜD — RECRUITMENT / HIRING TRACKER"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=TUV_BLUE_DARK)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 24

    # Summary strip (computed dynamically from whatever positions exist)
    total = len(position_rows)
    joined = sum(1 for r in position_rows if (r["status"] or "").strip().lower() == "joined")
    yet_to_join = sum(1 for r in position_rows if (r["status"] or "").strip().lower() == "yet to join")
    interviews = sum(1 for r in position_rows if "interview" in (r["status"] or "").lower())
    total_sf = sum(n for n in (_to_number(r["sf_applications"]) for r in position_rows) if n is not None)
    total_li_apps = sum(n for n in (_to_number(r["linkedin_applications"]) for r in position_rows) if n is not None)

    sum_labels = ["Total Positions", "Joined", "Yet to Join", "In Interview Process",
                  "Total SF Applications", "Total LinkedIn Applications"]
    for i, lbl in enumerate(sum_labels, 1):
        c = ws.cell(row=3, column=i, value=lbl)
        c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        c.fill = blue_fill; c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sum_vals = [total, joined, yet_to_join, interviews, int(total_sf), int(total_li_apps)]
    for i, v in enumerate(sum_vals, 1):
        c = ws.cell(row=4, column=i, value=v)
        c.font = Font(name="Arial", size=13, bold=True, color=TUV_BLUE_DARK)
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[4].height = 24

    # Table
    headers = ["S.No", "Position", "Location", "Status", "Date Opened", "Date Closed / DOJ",
               "No. of Days", "SF Applications", "LinkedIn Views", "LinkedIn Applications",
               "Recruiter Hiring Experience", "Sourcing Notes / Suggestions"]
    hr = 6
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=i, value=h)
        c.font = white; c.fill = blue_fill; c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[hr].height = 30

    for n, r in enumerate(position_rows):
        vals = [n + 1, r["position"], r["location"], r["status"], r["date_opened"], r["date_closed"],
                r["days"], r["sf_applications"], r["linkedin_views"], r["linkedin_applications"],
                r["recruiter_experience"], r["notes"]]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=hr + 1 + n, column=i, value=v)
            c.border = border
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(horizontal="left" if i in (2, 11, 12) else "center",
                                    vertical="center", wrap_text=(i in (2, 11, 12)))
            if n % 2 == 1:
                c.fill = grey_fill

    if not position_rows:
        c = ws.cell(row=hr + 1, column=1, value="No positions added yet.")
        ws.merge_cells(start_row=hr + 1, start_column=1, end_row=hr + 1, end_column=len(headers))
        c.font = Font(name="Arial", size=10, italic=True, color="6B7280")
        c.alignment = Alignment(horizontal="center", vertical="center")

    widths = [6, 22, 12, 20, 14, 22, 12, 14, 14, 16, 26, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[_col(i)].width = w
    ws.freeze_panes = f"A{hr+1}"


@app.route("/export")
def export_excel():
    filters = {
        "text": request.args.get("text", ""),
        "role": request.args.get("role", ""),
        "iso": request.args.get("iso", ""),
        "source": request.args.get("source", ""),
        "status": request.args.get("status", ""),
        "linkedin_sourced": request.args.get("linkedin_sourced", ""),
        "min_exp": request.args.get("min_exp", ""),
    }
    rows = query(filters)

    wb = Workbook()

    white = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    blue_fill = PatternFill("solid", fgColor=TUV_BLUE)
    grey_fill = PatternFill("solid", fgColor="F2F4F6")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ===== Sheet 1: Hiring Tracker (role-level summary) =====
    position_rows = query_positions()
    build_hiring_tracker(wb.active, white, blue_fill, grey_fill, border, position_rows)

    # ===== Sheet 2: Candidates =====
    ws = wb.create_sheet("Candidates")

    # Title banner (matches uploaded tracker layout)
    ws.merge_cells("A1:L1")
    ws["A1"] = "TÜV SÜD — CANDIDATE TRACKER / EXPORT"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=TUV_BLUE_DARK)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 24

    headers = ["S.No", "Candidate Name", "Role Applied For", "Location", "Experience (Yrs)",
               "Certifications", "ISO Certified", "Source", "LinkedIn Sourced",
               "Interview Stage", "Status", "Recruiter Notes"]
    header_row = 3
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = white; c.fill = blue_fill; c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 30

    for n, r in enumerate(rows, 1):
        vals = [n, r["name"], r["role"], r["location"], r["experience"],
                r["certifications"], r["iso_certified"], r["source"],
                r["linkedin_sourced"], r["stage"], r["status"], r["notes"]]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=n + header_row, column=i, value=v)
            c.border = border
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(horizontal="left" if i in (2, 6, 12) else "center",
                                    vertical="center", wrap_text=(i in (6, 12)))
            if n % 2 == 0:
                c.fill = grey_fill

    widths = [6, 22, 22, 12, 14, 26, 12, 12, 14, 18, 14, 28]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="TUV_SUD_Candidates.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    app.run(debug=True, port=5000)
